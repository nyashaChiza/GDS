
from django.http import HttpResponse
from transactions.models import Transaction
import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timedelta
from stock.models import Reciept
from requisition.models import Requisition


def generate_sales_report(month):

    # Parse the month to a datetime object
    month_start = datetime.strptime(month, "%Y-%m")
    month_end = (month_start + timedelta(days=31)).replace(day=1) - timedelta(days=1)

    # Query the Transaction model for the specified month
    transactions = Transaction.objects.filter(created__range=[month_start, month_end])

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Define headers
    headers = ["Transaction ID", "Customer", "Product", "Quantity", "Unit Cost", "Total Cost", "Date"]
    ws.append(headers)

    # Apply a bold font to the headers
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)

    # Populate with sales report data
    for transaction in transactions:
        row = [
            transaction.order_number,
            transaction.customer,
            transaction.product.name,
            transaction.quantity,
            transaction.unit_cost,
            transaction.quantity * transaction.unit_cost,
            transaction.created.strftime("%Y-%m-%d")
        ]
        ws.append(row)

    # Save the workbook to an HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=sales-report_{month}.xlsx'
    wb.save(response)

    return response

def generate_requisition_report(month):
    # Parse the month to a datetime object
    month_start = datetime.strptime(month, "%Y-%m")
    month_end = (month_start + timedelta(days=31)).replace(day=1) - timedelta(days=1)

    # Query the Requisition model for the specified month
    requisitions = Requisition.objects.filter(created__range=[month_start, month_end])

    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requisition Report"

    # Define headers
    headers = ["Requisition ID", "Type", "Description", "Status", "Date Created", "Date Updated"]
    ws.append(headers)

    # Apply a bold font to the headers
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)

    # Populate with requisition report data
    for requisition in requisitions:
        row = [
            requisition.id,
            requisition.requisition_type,
            requisition.description,
            requisition.status,
            requisition.created.strftime("%Y-%m-%d"),
            requisition.updated.strftime("%Y-%m-%d")
        ]
        ws.append(row)

    # Save the workbook to an HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=requisition-report_{month}.xlsx'
    wb.save(response)

    return response

def generate_inventory_report(month):
    # Parse the month to a datetime object
    month_start = datetime.strptime(month, "%Y-%m")
    month_end = (month_start + timedelta(days=31)).replace(day=1) - timedelta(days=1)

    # Query transactions for the given month
    transactions = Transaction.objects.filter(created__range=[month_start, month_end])

    # Query receipts for the given month
    receipts = Reciept.objects.filter(created__range=[month_start, month_end])

    # Initialize Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory Report'

    # Define headers for the report
    ws.append(['Product', 'Initial Inventory', 'Transactions Quantity', 'Receipts Quantity', 'Final Inventory'])

    # Initialize a dictionary to store inventory changes
    inventory_changes = {}

    # Calculate net inventory changes from transactions
    for transaction in transactions:
        product_id = transaction.product.id
        if product_id not in inventory_changes:
            inventory_changes[product_id] = {
                'product': str(transaction.product),
                'initial_inventory': 0,
                'transactions_quantity': 0,
                'receipts_quantity': 0,
                'final_inventory': 0
            }
        inventory_changes[product_id]['transactions_quantity'] += transaction.quantity

    # Calculate net inventory changes from receipts
    for receipt in receipts:
        product_id = receipt.stock.id
        if product_id not in inventory_changes:
            inventory_changes[product_id] = {
                'product': str(receipt.stock),
                'initial_inventory': 0,
                'transactions_quantity': 0,
                'receipts_quantity': 0,
                'final_inventory': 0
            }
        inventory_changes[product_id]['receipts_quantity'] += receipt.quantity

    # Calculate initial and final inventory for each product
    for product_id, changes in inventory_changes.items():
        initial_inventory = 0  # You can fetch this from your actual inventory data if available
        transactions_quantity = changes['transactions_quantity']
        receipts_quantity = changes['receipts_quantity']
        final_inventory = initial_inventory + receipts_quantity - transactions_quantity

        changes['initial_inventory'] = initial_inventory
        changes['final_inventory'] = final_inventory

        # Append the data to the Excel sheet
        ws.append([
            changes['product'],
            changes['initial_inventory'],
            changes['transactions_quantity'],
            changes['receipts_quantity'],
            changes['final_inventory']
        ])

    # Save the workbook to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=inventory-report_{month}.xlsx'
    wb.save(response)

    return response